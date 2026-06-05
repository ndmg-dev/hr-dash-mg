import openpyxl

wb = openpyxl.load_workbook('Ranking_Beneficios_Base_Dashboard.xlsx')
ws = wb['Base_Dashboard']

print('Dimensions:', ws.dimensions)
print('Max row:', ws.max_row, 'Max col:', ws.max_column)
print()

print('=== HEADERS (Row 1) ===')
headers = [cell.value for cell in ws[1]]
print(headers)
print()

print('=== FIRST 5 DATA ROWS ===')
for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
    print(list(row))
print()

print('=== ALL UNIQUE Companies ===')
companies = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1]:
        companies.add(row[1])
print(sorted(companies))
print()

print('=== ALL UNIQUE Benefits ===')
benefits = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[3]:
        benefits.add(row[3])
print(sorted(benefits))
print()

print('=== ALL UNIQUE Status_Oferta ===')
statuses = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[7]:
        statuses.add(row[7])
print(sorted(statuses))
print()

print('=== ALL UNIQUE Categoria_Beneficio ===')
cats = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[5]:
        cats.add(row[5])
print(sorted(cats))
print()

print('=== ALL UNIQUE Impacto_Atracao ===')
impacts = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[17]:
        impacts.add(row[17])
print(sorted(impacts))
print()

print('=== SAMPLE Valor_Monetario_BR values (first 10) ===')
vals = [row[8] for row in ws.iter_rows(min_row=2, values_only=True)]
print(vals[:10])
print()

print('=== SAMPLE Percentual_Cobertura values (first 10) ===')
pcts = [row[9] for row in ws.iter_rows(min_row=2, values_only=True)]
print(pcts[:10])
print()

print('=== Flag_Mendonca_Galvao values ===')
flags = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    flags.add(row[14])
print(flags)
print()

print('=== Analise_Mendonca_Galvao values ===')
vals_mg = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[15]:
        vals_mg.add(row[15])
print(sorted(vals_mg))
print()

print('=== Analise_Concorrentes values ===')
vals_conc = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[16]:
        vals_conc.add(row[16])
print(sorted(vals_conc))
print()

print('=== FULL DATA DUMP ===')
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    print(f'Row {i+2}:', list(row))
